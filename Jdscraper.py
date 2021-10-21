from selenium import webdriver
import pandas as pd
import time
from openpyxl import load_workbook

remain_city = []
def all_city():
    global remain_city
    from bs4 import BeautifulSoup
    import requests
    url = "https://en.wikipedia.org/wiki/List_of_cities_in_India_by_population"
    html = requests.get(url)

    soup = BeautifulSoup(html.text, 'html.parser')
    tables = soup.find_all("table")

    tables = tables[0]
    city_names = []
    # print(tables)
    rows = tables.find_all('tr')
    express = ['[', ']']
    for row in rows:
        cell = row.find_all('td')

        if len(cell) > 1:
            name = cell[1].text.strip()
            name = ''.join([i for i in name if not i.isdigit()])
            new_name = ""
            for i in name:
                if i not in express:
                    new_name += ''.join(i)
            city_names.append(new_name)

    for i in range(len(city_names)):
        remain_city.append(city_names[i])


def main(city, sno):
    PATH = 'C:\\Program Files (x86)\\chromedriver.exe'
    # op = webdriver.ChromeOptions()
    # op.add_argument('headless')
    driver = webdriver.Chrome(PATH)
    # driver.minimize_window()

    # print(url, sno)

    ### Enter the URL from JUSTDIAL here
    driver.get(f"https://www.justdial.com/{city}/Cbse-Schools/nct-10083838")
    nameee = "mainn" + str(sno)

    driver.execute_script("window.scrollTo(0, 1500)")
    time.sleep(2)
    driver.execute_script("window.scrollTo(0, 3000)")
    time.sleep(2)
    driver.execute_script("window.scrollTo(0, 4000)")
    time.sleep(2)

    driver.execute_script("window.scrollTo(0, 5000)")

    driver.execute_script("window.scrollTo(0, 6000)")
    time.sleep(2)

    driver.execute_script("window.scrollTo(0, 7000)")
    time.sleep(1)

    driver.execute_script("window.scrollTo(0, 9000)")

    driver.execute_script("window.scrollTo(0, 20000)")
    time.sleep(2)

    storeDetails = driver.find_elements_by_class_name('store-details')

    driver.quit()

    def toExcel(df):
        

        try:
            book = load_workbook('JDscrap.xlsx')
        except:
            with open('JDscrap.xlsx', 'w') as fp:
                pass
            book = load_workbook('JDscrap.xlsx')

        with pd.ExcelWriter('JDscrap.xlsx', mode='a') as writer:
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            print(writer.sheets)
            df.to_excel(writer, sheet_name=nameee, header=False, index=False)

    def strings_to_num(argument):
        switcher = {
            'dc': '+',
            'fe': '(',
            'hg': ')',
            'ba': '-',
            'acb': '0',
            'yz': '1',
            'wx': '2',
            'vu': '3',
            'ts': '4',
            'rq': '5',
            'po': '6',
            'nm': '7',
            'lk': '8',
            'ji': '9'
        }

        return switcher.get(argument, "nothing")

    nameList = []
    addressList = []
    numbersList = []
    # print(storeDetails)
    for i in range(len(storeDetails)):

        name = storeDetails[i].find_element_by_class_name('lng_cont_name').text
        address = storeDetails[i].find_element_by_class_name('cont_fl_addr').get_attribute('innerHTML')
        contactList = storeDetails[i].find_elements_by_class_name('mobilesv')

        myList = []

        for j in range(len(contactList)):
            myString = contactList[j].get_attribute('class').split("-")[1]

            myList.append(strings_to_num(myString))

        nameList.append(name)
        addressList.append(address)
        numbersList.append("".join(myList))

    # intialise data of lists.
    data = {'Name': nameList,
            'Address': addressList,
            'Phone': numbersList}

    # Create DataFrame
    df = pd.DataFrame(data)

    print(df)

    # df.to_excel('teacher.xlsx', header= False, index = False, mode = "a")
    toExcel(df)
    driver.quit()

all_city()
print(remain_city)
for i in range(len(remain_city)):
    try:
        main(remain_city[i], i)
        print(i)
    except:
        pass
